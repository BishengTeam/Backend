"""Batch-level, independently extractable export ZIP generation."""

from __future__ import annotations

import asyncio
import hashlib
import logging
import re
import shutil
import tempfile
import uuid
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Sequence

from sqlalchemy import delete, exists, func, select, update

from app.adapter.database import get_db_ctx
from app.domain.order.src.index import Order
from app.domain.plan.src.index import Plan
from app.domain.renshe.src.index import (
    EXPORTABLE_APPLICATION_STATUSES,
    MATERIAL_SPECS,
    MAX_EXPORT_VOLUME_BYTES,
    RensheApplication,
    RensheApplicationVersion,
    RensheAuditLog,
    RensheCleanupRun,
    RensheExportItem,
    RensheExportJob,
    RensheExportVolume,
    RensheMaterial,
)
from app.integrations.renshe_storage import RensheObjectStorage
from app.integrations.renshe_storage import assert_owned_source_key
from app.domain.user.src.index import UserRealname, UserStudent
from app.port.config import settings
from app.port.exceptions import (
    BusinessException,
    ConflictException,
    NotFoundException,
    ThirdPartyException,
)
from app.utils.audit import redact_sensitive_text
from app.schemas.renshe import (
    RensheExportJobResponse,
    RensheExportVolumeResponse,
    RensheSignedUrlResponse,
)


logger = logging.getLogger(__name__)

REGISTRATION_TEMPLATE_NAME = "报名信息.xlsx"
WORK_HISTORY_TEMPLATE_NAME = "工作经历.xlsx"
EXPORT_FIXED_OVERHEAD_BYTES = 4 * 1024 * 1024
EXPORT_ENTRY_OVERHEAD_BYTES = 1024
# Workbook rows and ZIP directory entries are small compared with materials,
# but reserving a bounded amount per candidate keeps the 10 GiB partition
# decision conservative even when a batch is close to the ceiling.
EXPORT_CANDIDATE_OVERHEAD_BYTES = 64 * 1024
# The first release accepts only related-major current graduating students.
# Qualification is intentionally a human-review decision; this value is the
# frozen label written to the official workbook, not an automated eligibility
# predicate.
APPLICATION_CONDITION = "相关专业的在读应届学生"
EDUCATION_LABELS = {
    "secondary_vocational": "中专",
    "associate": "大专",
    "bachelor": "大学本科",
    "master": "硕士",
    "doctorate": "博士",
}
MATERIAL_EXPORT_NAMES = {
    "id_card_front": "证件正面.jpg",
    "id_card_back": "证件背面.jpg",
    "portrait": "二寸免冠照片.jpg",
    "student_card": "学生证.jpg",
    "xuexin_registration": "学信网电子注册表.pdf",
    "education_proof": "学历证明.jpg",
}


@dataclass(frozen=True, slots=True)
class ExportMaterial:
    id: int
    kind: str
    storage_key: str
    size_bytes: int
    sha256: str


@dataclass(frozen=True, slots=True)
class ExportCandidate:
    application_id: int
    version_id: int
    estimated_size_bytes: int
    realname_snapshot: dict
    student_snapshot: dict
    form_data: dict
    materials: tuple[ExportMaterial, ...]


def partition_export_candidates(
    candidates: Sequence[tuple[int, int]],
    *,
    max_volume_bytes: int = MAX_EXPORT_VOLUME_BYTES,
    fixed_overhead_bytes: int = EXPORT_FIXED_OVERHEAD_BYTES,
    per_candidate_overhead_bytes: int = 0,
) -> list[list[int]]:
    """First-fit candidates in stable order without splitting a candidate.

    ``candidates`` contains ``(application_id, estimated_material_bytes)``.
    Estimates deliberately use stored/uncompressed sizes because generated ZIPs
    use ``ZIP_STORED``; this makes the 10 GiB ceiling predictable.  Callers can
    reserve additional per-candidate workbook/ZIP overhead without changing
    the no-split guarantee.
    """

    if max_volume_bytes <= fixed_overhead_bytes:
        raise ValueError("分卷上限必须大于固定文件开销")
    volumes: list[list[int]] = []
    current: list[int] = []
    current_size = fixed_overhead_bytes
    for application_id, estimated_size in candidates:
        candidate_size = (
            max(0, estimated_size)
            + (len(MATERIAL_SPECS) * EXPORT_ENTRY_OVERHEAD_BYTES)
            + max(0, per_candidate_overhead_bytes)
        )
        if candidate_size + fixed_overhead_bytes > max_volume_bytes:
            raise ValueError(f"报名 {application_id} 的材料无法放入单个分卷")
        if current and current_size + candidate_size > max_volume_bytes:
            volumes.append(current)
            current = []
            current_size = fixed_overhead_bytes
        current.append(application_id)
        current_size += candidate_size
    if current:
        volumes.append(current)
    return volumes


def build_registration_row(candidate: ExportCandidate, plan: Plan) -> list[object | None]:
    """Map one immutable submission snapshot to the official 25 columns."""

    realname = candidate.realname_snapshot
    student = candidate.student_snapshot
    form = candidate.form_data
    return [
        realname.get("real_name"),
        realname.get("gender"),
        realname.get("birth_date"),
        "居民身份证",
        realname.get("id_card_number"),
        "院校学生",
        EDUCATION_LABELS.get(student.get("education"), student.get("education")),
        realname.get("ethnicity"),
        realname.get("political_status"),
        form.get("contact_phone"),
        form.get("mailing_address"),
        None,
        "应届毕业生",
        plan.occupation_name or "信息安全管理员",
        "四级",
        None,
        None,
        student.get("enrollment_date"),
        None,
        plan.occupation_code
        or "4-04-04-02-网络与信息安全管理员-信息安全管理员-中级工",
        APPLICATION_CONDITION,
        plan.application_type or "学历型",
        plan.exam_type or "新考",
        None,
        form.get("email"),
    ]


class RensheExportService:
    def __init__(self, storage: RensheObjectStorage | None = None) -> None:
        self.storage = storage or RensheObjectStorage()

    @staticmethod
    def _now() -> datetime:
        return datetime.now(timezone.utc)

    async def create_job(self, *, plan_id: int, admin_id: int) -> RensheExportJobResponse:
        async with get_db_ctx() as db:
            plan = await db.scalar(
                select(Plan).where(Plan.id == plan_id).with_for_update()
            )
            if plan is None or plan.product_type != "RS-ZY":
                raise NotFoundException("人社报名批次")
            if plan.status not in {"published", "registration_closed", "finalized"}:
                raise ConflictException("当前批次状态不能导出")
            await self._assert_export_window_open(db, plan)

            # A double click or two browser tabs must not enqueue unbounded
            # concurrent generations for the same batch.  Once the active job
            # finishes, an explicit new request creates the next generation.
            active_job = await db.scalar(
                select(RensheExportJob)
                .where(
                    RensheExportJob.plan_id == plan_id,
                    RensheExportJob.status.in_(("queued", "running")),
                )
                .order_by(RensheExportJob.generation_no.desc())
                .with_for_update()
                .limit(1)
            )
            if active_job is not None:
                volumes = (
                    await db.execute(
                        select(RensheExportVolume)
                        .where(RensheExportVolume.job_id == active_job.id)
                        .order_by(RensheExportVolume.volume_no)
                    )
                ).scalars().all()
                return self._job_response(active_job, volumes)

            candidates = await self._eligible_candidate_rows(db, plan_id)
            if not candidates:
                raise BusinessException("当前批次没有可导出的考生")
            generation_no = (
                await db.scalar(
                    select(func.max(RensheExportJob.generation_no)).where(
                        RensheExportJob.plan_id == plan_id
                    )
                )
                or 0
            ) + 1
            job = RensheExportJob(
                plan_id=plan_id,
                generation_no=generation_no,
                requested_by_admin_id=admin_id,
                status="queued",
                candidate_total=len(candidates),
            )
            db.add(job)
            await db.flush()
            for application_id, version_id, estimated_size in candidates:
                db.add(
                    RensheExportItem(
                        job_id=job.id,
                        application_id=application_id,
                        version_id=version_id,
                        status="queued",
                        estimated_size_bytes=estimated_size,
                    )
                )
            db.add(
                RensheAuditLog(
                    actor_type="admin",
                    actor_id=admin_id,
                    action="export.request",
                    object_type="export_job",
                    object_id=job.id,
                    result="succeeded",
                    summary={
                        "plan_id": plan_id,
                        "generation_no": generation_no,
                        "candidate_total": len(candidates),
                    },
                )
            )
            await db.commit()
            await db.refresh(job)
            return self._job_response(job, [])

    async def get_job(self, job_id: int) -> RensheExportJobResponse:
        async with get_db_ctx() as db:
            job = await db.get(RensheExportJob, job_id)
            if job is None:
                raise NotFoundException("人社导出任务")
            volumes = (
                await db.execute(
                    select(RensheExportVolume)
                    .where(RensheExportVolume.job_id == job.id)
                    .order_by(RensheExportVolume.volume_no)
                )
            ).scalars().all()
            return self._job_response(job, volumes)

    async def list_jobs(self, plan_id: int) -> list[RensheExportJobResponse]:
        async with get_db_ctx() as db:
            jobs = (
                await db.execute(
                    select(RensheExportJob)
                    .where(RensheExportJob.plan_id == plan_id)
                    .order_by(RensheExportJob.generation_no.desc())
                )
            ).scalars().all()
            if not jobs:
                return []
            volumes = (
                await db.execute(
                    select(RensheExportVolume)
                    .where(RensheExportVolume.job_id.in_([job.id for job in jobs]))
                    .order_by(RensheExportVolume.job_id, RensheExportVolume.volume_no)
                )
            ).scalars().all()
            by_job: dict[int, list[RensheExportVolume]] = {}
            for volume in volumes:
                by_job.setdefault(volume.job_id, []).append(volume)
            return [self._job_response(job, by_job.get(job.id, [])) for job in jobs]

    async def retry_job(self, *, job_id: int, admin_id: int) -> RensheExportJobResponse:
        async with get_db_ctx() as db:
            plan_id = await db.scalar(
                select(RensheExportJob.plan_id).where(RensheExportJob.id == job_id)
            )
            if plan_id is None:
                raise NotFoundException("人社导出任务")
            plan = await db.scalar(
                select(Plan).where(Plan.id == plan_id).with_for_update()
            )
            if plan is None or plan.product_type != "RS-ZY":
                raise NotFoundException("人社报名批次")
            await self._assert_export_window_open(db, plan)
            job = await db.scalar(
                select(RensheExportJob)
                .where(RensheExportJob.id == job_id)
                .with_for_update()
            )
            if job is None:
                raise NotFoundException("人社导出任务")
            if job.status != "failed":
                raise ConflictException("只有失败的导出任务可以重试")
            job.status = "queued"
            job.candidate_processed = 0
            job.volume_count = 0
            job.heartbeat_at = None
            job.started_at = None
            job.finished_at = None
            job.retry_count += 1
            job.last_error = None
            db.add(
                RensheAuditLog(
                    actor_type="admin",
                    actor_id=admin_id,
                    action="export.retry",
                    object_type="export_job",
                    object_id=job.id,
                    result="succeeded",
                    summary={"retry_count": job.retry_count},
                )
            )
            await db.commit()
        return await self.get_job(job_id)

    async def _assert_export_window_open(self, db, plan: Plan) -> None:
        if plan.cleanup_due_at is not None and plan.cleanup_due_at <= self._now():
            raise ConflictException("批次材料已到清理期限，不能新建或重试导出")
        cleanup_status = await db.scalar(
            select(RensheCleanupRun.status)
            .where(RensheCleanupRun.plan_id == plan.id)
            .order_by(RensheCleanupRun.id.desc())
            .limit(1)
        )
        if cleanup_status in {"running", "succeeded"}:
            raise ConflictException("批次材料正在清理或已清理，不能新建或重试导出")

    async def material_signed_url(
        self,
        *,
        material_id: int,
        admin_id: int,
        download: bool,
        ip_address: str | None,
    ) -> RensheSignedUrlResponse:
        action = "material.download" if download else "material.preview"
        application_id = None
        version_id = None
        kind = "unknown"
        try:
            async with get_db_ctx() as db:
                material = await db.get(RensheMaterial, material_id)
                if material is None or material.is_deleted:
                    raise NotFoundException("人社报名材料")
                storage_key = material.storage_key
                application_id = material.application_id
                version_id = material.version_id
                kind = material.kind
        except Exception as exc:
            await self._access_audit(
                actor_type="admin",
                actor_id=admin_id,
                action=action,
                object_type="material",
                object_id=material_id,
                application_id=application_id,
                version_id=version_id,
                material_id=material_id,
                ip_address=ip_address,
                result="failed",
                summary={"kind": kind, "error_type": type(exc).__name__},
            )
            raise
        try:
            filename = f"{kind}{Path(storage_key).suffix.lower()}" if download else None
            url = await self.storage.signed_get_url(
                storage_key, download_filename=filename
            )
        except Exception as exc:
            await self._access_audit(
                actor_type="admin",
                actor_id=admin_id,
                action=action,
                object_type="material",
                object_id=material_id,
                application_id=application_id,
                version_id=version_id,
                material_id=material_id,
                ip_address=ip_address,
                result="failed",
                summary={"error_type": type(exc).__name__},
            )
            raise
        await self._access_audit(
            actor_type="admin",
            actor_id=admin_id,
            action=action,
            object_type="material",
            object_id=material_id,
            application_id=application_id,
            version_id=version_id,
            material_id=material_id,
            ip_address=ip_address,
            result="succeeded",
            summary={"kind": kind},
        )
        return RensheSignedUrlResponse(
            url=url, expires_in=settings.ALIYUN_OSS_SIGNED_URL_TTL_SECONDS
        )

    async def verification_material_signed_url(
        self,
        *,
        user_id: int,
        kind: str,
        admin_id: int,
        download: bool,
        ip_address: str | None,
    ) -> RensheSignedUrlResponse:
        return await self._verification_material_signed_url(
            user_id=user_id,
            kind=kind,
            actor_type="admin",
            actor_id=admin_id,
            download=download,
            ip_address=ip_address,
        )

    async def user_verification_material_signed_url(
        self,
        *,
        user_id: int,
        kind: str,
        download: bool,
        ip_address: str | None,
    ) -> RensheSignedUrlResponse:
        return await self._verification_material_signed_url(
            user_id=user_id,
            kind=kind,
            actor_type="user",
            actor_id=user_id,
            download=download,
            ip_address=ip_address,
        )

    async def _verification_material_signed_url(
        self,
        *,
        user_id: int,
        kind: str,
        actor_type: str,
        actor_id: int,
        download: bool,
        ip_address: str | None,
    ) -> RensheSignedUrlResponse:
        action = (
            "verification_material.download"
            if download
            else "verification_material.preview"
        )
        try:
            async with get_db_ctx() as db:
                realname = await db.scalar(
                    select(UserRealname).where(UserRealname.user_id == user_id)
                )
                student = await db.scalar(
                    select(UserStudent).where(UserStudent.user_id == user_id)
                )
                source_keys = {
                    "id_card_front": realname.id_card_front_oss if realname else None,
                    "id_card_back": realname.id_card_back_oss if realname else None,
                    "portrait": realname.avatar_oss if realname else None,
                    "student_card": student.student_card_oss if student else None,
                    "xuexin_registration": student.enrollment_pdf_oss if student else None,
                    "education_proof": student.degree_cert_oss if student else None,
                }
                if kind not in source_keys or not source_keys[kind]:
                    raise NotFoundException("用户认证材料")
                storage_key = source_keys[kind]
                assert_owned_source_key(user_id, storage_key)
        except Exception as exc:
            await self._access_audit(
                actor_type=actor_type,
                actor_id=actor_id,
                action=action,
                object_type="verification_material",
                object_id=user_id,
                application_id=None,
                version_id=None,
                material_id=None,
                ip_address=ip_address,
                result="failed",
                summary={"kind": kind, "error_type": type(exc).__name__},
            )
            raise
        try:
            filename = f"{kind}{Path(storage_key).suffix.lower()}" if download else None
            url = await self.storage.signed_get_url(
                storage_key, download_filename=filename
            )
        except Exception as exc:
            await self._access_audit(
                actor_type=actor_type,
                actor_id=actor_id,
                action=action,
                object_type="verification_material",
                object_id=user_id,
                application_id=None,
                version_id=None,
                material_id=None,
                ip_address=ip_address,
                result="failed",
                summary={"kind": kind, "error_type": type(exc).__name__},
            )
            raise
        await self._access_audit(
            actor_type=actor_type,
            actor_id=actor_id,
            action=action,
            object_type="verification_material",
            object_id=user_id,
            application_id=None,
            version_id=None,
            material_id=None,
            ip_address=ip_address,
            result="succeeded",
            summary={"kind": kind},
        )
        return RensheSignedUrlResponse(
            url=url, expires_in=settings.ALIYUN_OSS_SIGNED_URL_TTL_SECONDS
        )

    async def volume_signed_url(
        self,
        *,
        volume_id: int,
        admin_id: int,
        ip_address: str | None,
    ) -> RensheSignedUrlResponse:
        job_id: int | None = None
        volume_no: int | None = None
        try:
            async with get_db_ctx() as db:
                row = (
                    await db.execute(
                        select(RensheExportVolume, RensheExportJob)
                        .join(
                            RensheExportJob,
                            RensheExportJob.id == RensheExportVolume.job_id,
                        )
                        .where(RensheExportVolume.id == volume_id)
                    )
                ).one_or_none()
                if row is None:
                    raise NotFoundException("人社导出分卷")
                volume, job = row
                job_id = job.id
                volume_no = volume.volume_no
                if volume.status != "succeeded" or not volume.storage_key:
                    raise ConflictException("导出分卷尚不可下载或已清理")
                storage_key = volume.storage_key
                filename = f"人社报名_批次{job.plan_id}_第{volume.volume_no}卷.zip"
        except Exception as exc:
            await self._access_audit(
                actor_type="admin",
                actor_id=admin_id,
                action="export.download",
                object_type="export_volume",
                object_id=volume_id,
                application_id=None,
                version_id=None,
                material_id=None,
                ip_address=ip_address,
                result="failed",
                summary={
                    "job_id": job_id,
                    "volume_no": volume_no,
                    "error_type": type(exc).__name__,
                },
            )
            raise
        try:
            url = await self.storage.signed_get_url(
                storage_key, download_filename=filename
            )
        except Exception as exc:
            await self._access_audit(
                actor_type="admin",
                actor_id=admin_id,
                action="export.download",
                object_type="export_volume",
                object_id=volume_id,
                application_id=None,
                version_id=None,
                material_id=None,
                ip_address=ip_address,
                result="failed",
                summary={"job_id": job_id, "error_type": type(exc).__name__},
            )
            raise
        await self._access_audit(
            actor_type="admin",
            actor_id=admin_id,
            action="export.download",
            object_type="export_volume",
            object_id=volume_id,
            application_id=None,
            version_id=None,
            material_id=None,
            ip_address=ip_address,
            result="succeeded",
            summary={"job_id": job_id, "volume_no": volume_no},
        )
        return RensheSignedUrlResponse(
            url=url, expires_in=settings.ALIYUN_OSS_SIGNED_URL_TTL_SECONDS
        )

    async def process_next_job(self) -> bool:
        job_id = await self._claim_next_job()
        if job_id is None:
            return False
        heartbeat = asyncio.create_task(self._heartbeat_loop(job_id))
        try:
            await self._run_claimed_job(job_id)
        except Exception as exc:
            logger.error(
                "human-resources export job failed: job_id=%s exception_type=%s",
                job_id,
                type(exc).__name__,
            )
            await self._mark_failed(job_id, exc)
        finally:
            heartbeat.cancel()
            try:
                await heartbeat
            except asyncio.CancelledError:
                pass
        return True

    async def _claim_next_job(self) -> int | None:
        await self._recover_stale_jobs()
        async with get_db_ctx() as db:
            # Do not let a queued job whose batch is already being cleaned
            # block every later batch.  We inspect the queue in stable pages;
            # each candidate is re-locked after its batch lock so concurrent
            # workers/finalizers remain safe.  A bounded page avoids loading a
            # pathological backlog into one worker transaction while the loop
            # still guarantees progress when early batches are blocked.
            page_size = 100
            inspected_ids: set[int] = set()
            while True:
                queue_stmt = select(RensheExportJob).where(
                    RensheExportJob.status == "queued"
                )
                if inspected_ids:
                    queue_stmt = queue_stmt.where(
                        ~RensheExportJob.id.in_(inspected_ids)
                    )
                result = await db.execute(
                    queue_stmt
                    .order_by(RensheExportJob.id)
                    .limit(page_size)
                )
                candidates = result.scalars().all()
                if not candidates:
                    return None
                for candidate in candidates:
                    inspected_ids.add(candidate.id)
                    # Serialize the worker claim with batch
                    # finalization/cleanup.  The cleanup worker takes the
                    # same plan lock before marking a run running, so a
                    # queued export cannot slip in after cleanup has claimed
                    # the batch.
                    plan = await db.scalar(
                        select(Plan)
                        .where(Plan.id == candidate.plan_id)
                        .with_for_update()
                    )
                    if plan is None:
                        continue
                    cleanup_status = await db.scalar(
                        select(RensheCleanupRun.status)
                        .where(RensheCleanupRun.plan_id == plan.id)
                        .order_by(RensheCleanupRun.id.desc())
                        .limit(1)
                    )
                    if cleanup_status in {"running", "succeeded"}:
                        continue
                    job = await db.scalar(
                        select(RensheExportJob)
                        .where(
                            RensheExportJob.id == candidate.id,
                            RensheExportJob.status == "queued",
                        )
                        .with_for_update(skip_locked=True)
                    )
                    if job is None:
                        continue
                    now = self._now()
                    job.status = "running"
                    job.started_at = now
                    job.heartbeat_at = now
                    job.finished_at = None
                    job.last_error = None
                    await db.commit()
                    return job.id

    async def _run_claimed_job(self, job_id: int) -> None:
        template_registration, template_work = self._template_paths()
        old_keys: list[str] = []
        async with get_db_ctx() as db:
            old_keys = list(
                (
                    await db.execute(
                        select(RensheExportVolume.storage_key).where(
                            RensheExportVolume.job_id == job_id,
                            RensheExportVolume.storage_key.is_not(None),
                        )
                    )
                ).scalars().all()
            )
        if old_keys:
            await self.storage.delete_many(old_keys)

        async with get_db_ctx() as db:
            job = await db.scalar(
                select(RensheExportJob)
                .where(RensheExportJob.id == job_id, RensheExportJob.status == "running")
                .with_for_update()
            )
            if job is None:
                raise ConflictException("导出任务未处于运行状态")
            plan = await db.get(Plan, job.plan_id)
            if plan is None:
                raise NotFoundException("人社报名批次")
            await db.execute(
                update(RensheExportItem)
                .where(RensheExportItem.job_id == job.id)
                .values(volume_id=None)
            )
            await db.execute(delete(RensheExportVolume).where(RensheExportVolume.job_id == job.id))
            await db.execute(delete(RensheExportItem).where(RensheExportItem.job_id == job.id))
            eligible = await self._eligible_candidate_rows(db, job.plan_id)
            if not eligible:
                raise BusinessException("当前批次没有可导出的考生")
            for application_id, version_id, estimated_size in eligible:
                db.add(
                    RensheExportItem(
                        job_id=job.id,
                        application_id=application_id,
                        version_id=version_id,
                        status="queued",
                        estimated_size_bytes=estimated_size,
                    )
                )
            job.candidate_total = len(eligible)
            job.candidate_processed = 0
            job.volume_count = 0
            job.heartbeat_at = self._now()
            await db.commit()

        plan, candidates = await self._load_candidates(job_id)
        fixed_overhead = (
            template_registration.stat().st_size
            + template_work.stat().st_size
            + EXPORT_FIXED_OVERHEAD_BYTES
        )
        partitions = partition_export_candidates(
            [(candidate.application_id, candidate.estimated_size_bytes) for candidate in candidates],
            fixed_overhead_bytes=fixed_overhead,
            per_candidate_overhead_bytes=EXPORT_CANDIDATE_OVERHEAD_BYTES,
        )
        by_application = {candidate.application_id: candidate for candidate in candidates}
        for volume_no, application_ids in enumerate(partitions, start=1):
            volume_candidates = [by_application[item] for item in application_ids]
            volume_id, storage_key = await self._create_volume(
                job_id=job_id,
                plan_id=plan.id,
                generation_no=await self._job_generation(job_id),
                volume_no=volume_no,
                candidate_count=len(volume_candidates),
            )
            size_bytes, sha256 = await self._generate_volume(
                plan=plan,
                job_id=job_id,
                volume_no=volume_no,
                storage_key=storage_key,
                candidates=volume_candidates,
                registration_template=template_registration,
                work_template=template_work,
            )
            await self._complete_volume(
                job_id=job_id,
                volume_id=volume_id,
                application_ids=application_ids,
                size_bytes=size_bytes,
                sha256=sha256,
            )

        async with get_db_ctx() as db:
            job = await db.scalar(
                select(RensheExportJob)
                .where(RensheExportJob.id == job_id)
                .with_for_update()
            )
            if job is None:
                raise NotFoundException("人社导出任务")
            now = self._now()
            job.status = "succeeded"
            job.finished_at = now
            job.heartbeat_at = now
            job.volume_count = len(partitions)
            db.add(
                RensheAuditLog(
                    actor_type="system",
                    actor_id=None,
                    action="export.generate",
                    object_type="export_job",
                    object_id=job.id,
                    result="succeeded",
                    summary={
                        "candidate_total": job.candidate_total,
                        "volume_count": len(partitions),
                    },
                )
            )
            await db.commit()

    async def _generate_volume(
        self,
        *,
        plan: Plan,
        job_id: int,
        volume_no: int,
        storage_key: str,
        candidates: list[ExportCandidate],
        registration_template: Path,
        work_template: Path,
    ) -> tuple[int, str]:
        with tempfile.TemporaryDirectory(prefix=f"renshe-export-{job_id}-{volume_no}-") as raw:
            temp_root = Path(raw)
            registration_path = temp_root / REGISTRATION_TEMPLATE_NAME
            work_path = temp_root / WORK_HISTORY_TEMPLATE_NAME
            zip_path = temp_root / f"volume-{volume_no}.zip"
            await asyncio.to_thread(
                self._write_registration_workbook,
                registration_template,
                registration_path,
                candidates,
                plan,
            )
            await asyncio.to_thread(shutil.copyfile, work_template, work_path)

            archive = zipfile.ZipFile(
                zip_path,
                mode="w",
                compression=zipfile.ZIP_STORED,
                allowZip64=True,
            )
            try:
                await asyncio.to_thread(
                    archive.write, registration_path, REGISTRATION_TEMPLATE_NAME
                )
                await asyncio.to_thread(
                    archive.write, work_path, WORK_HISTORY_TEMPLATE_NAME
                )
                objects_dir = temp_root / "objects"
                objects_dir.mkdir()
                for candidate in candidates:
                    id_card = str(candidate.realname_snapshot.get("id_card_number") or "")
                    if not re.fullmatch(r"\d{17}[\dXx]", id_card):
                        raise BusinessException("导出快照中的身份证号无效")
                    for material in candidate.materials:
                        local_path = objects_dir / uuid.uuid4().hex
                        await self.storage.download_file(material.storage_key, local_path)
                        actual_size, actual_sha256 = await asyncio.to_thread(
                            self._file_size_and_sha256, local_path
                        )
                        if actual_size != material.size_bytes or actual_sha256 != material.sha256:
                            raise ConflictException("导出材料完整性校验失败")
                        export_name = MATERIAL_EXPORT_NAMES[material.kind]
                        archive_name = f"材料/{id_card}/{id_card}_{export_name}"
                        await asyncio.to_thread(archive.write, local_path, archive_name)
                        await asyncio.to_thread(local_path.unlink)
            finally:
                await asyncio.to_thread(archive.close)

            size_bytes, sha256 = await asyncio.to_thread(
                self._file_size_and_sha256, zip_path
            )
            if size_bytes > MAX_EXPORT_VOLUME_BYTES:
                raise BusinessException("生成分卷超过 10 GiB 上限")
            await self.storage.upload_file(storage_key, zip_path, "application/zip")
            return size_bytes, sha256

    @staticmethod
    def _write_registration_workbook(
        template_path: Path,
        output_path: Path,
        candidates: list[ExportCandidate],
        plan: Plan,
    ) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ThirdPartyException("缺少 Excel 模板处理依赖 openpyxl") from exc
        workbook = load_workbook(template_path)
        try:
            worksheet = workbook["Sheet1"]
            for row_no, candidate in enumerate(candidates, start=2):
                values = build_registration_row(candidate, plan)
                for column_no, value in enumerate(values, start=1):
                    worksheet.cell(row=row_no, column=column_no, value=value)
            workbook.save(output_path)
        finally:
            workbook.close()

    async def _load_candidates(self, job_id: int) -> tuple[Plan, list[ExportCandidate]]:
        async with get_db_ctx() as db:
            job = await db.get(RensheExportJob, job_id)
            if job is None:
                raise NotFoundException("人社导出任务")
            plan = await db.get(Plan, job.plan_id)
            if plan is None:
                raise NotFoundException("人社报名批次")
            rows = (
                await db.execute(
                    select(RensheExportItem, RensheApplicationVersion)
                    .join(
                        RensheApplicationVersion,
                        RensheApplicationVersion.id == RensheExportItem.version_id,
                    )
                    .where(RensheExportItem.job_id == job_id)
                    .order_by(RensheExportItem.application_id)
                )
            ).all()
            version_ids = [version.id for _, version in rows]
            materials = (
                await db.execute(
                    select(RensheMaterial)
                    .where(
                        RensheMaterial.version_id.in_(version_ids),
                        RensheMaterial.is_deleted.is_(False),
                    )
                    .order_by(RensheMaterial.version_id, RensheMaterial.kind)
                )
            ).scalars().all()
            by_version: dict[int, list[RensheMaterial]] = {}
            for material in materials:
                by_version.setdefault(material.version_id, []).append(material)
            candidates: list[ExportCandidate] = []
            expected_kinds = set(MATERIAL_SPECS)
            for item, version in rows:
                version_materials = by_version.get(version.id, [])
                if {material.kind for material in version_materials} != expected_kinds:
                    raise ConflictException("导出版本材料不完整")
                if not version.realname_snapshot or not version.student_snapshot:
                    raise ConflictException("导出版本敏感快照已清理")
                candidates.append(
                    ExportCandidate(
                        application_id=item.application_id,
                        version_id=version.id,
                        estimated_size_bytes=item.estimated_size_bytes or 0,
                        realname_snapshot=dict(version.realname_snapshot),
                        student_snapshot=dict(version.student_snapshot),
                        form_data=dict(version.form_data),
                        materials=tuple(
                            ExportMaterial(
                                id=material.id,
                                kind=material.kind,
                                storage_key=material.storage_key,
                                size_bytes=material.size_bytes,
                                sha256=material.sha256,
                            )
                            for material in version_materials
                        ),
                    )
                )
            return plan, candidates

    async def _create_volume(
        self,
        *,
        job_id: int,
        plan_id: int,
        generation_no: int,
        volume_no: int,
        candidate_count: int,
    ) -> tuple[int, str]:
        storage_key = (
            f"{settings.ALIYUN_OSS_PREFIX.strip('/') or 'renshe'}/exports/"
            f"{plan_id}/g{generation_no}/{uuid.uuid4().hex}.zip"
        )
        async with get_db_ctx() as db:
            volume = RensheExportVolume(
                job_id=job_id,
                volume_no=volume_no,
                status="running",
                candidate_count=candidate_count,
                storage_key=storage_key,
            )
            db.add(volume)
            await db.flush()
            await db.commit()
            return volume.id, storage_key

    async def _complete_volume(
        self,
        *,
        job_id: int,
        volume_id: int,
        application_ids: list[int],
        size_bytes: int,
        sha256: str,
    ) -> None:
        async with get_db_ctx() as db:
            volume = await db.scalar(
                select(RensheExportVolume)
                .where(RensheExportVolume.id == volume_id)
                .with_for_update()
            )
            job = await db.scalar(
                select(RensheExportJob)
                .where(RensheExportJob.id == job_id)
                .with_for_update()
            )
            if volume is None or job is None:
                raise NotFoundException("人社导出任务")
            now = self._now()
            volume.status = "succeeded"
            volume.size_bytes = size_bytes
            volume.sha256 = sha256
            volume.finished_at = now
            await db.execute(
                update(RensheExportItem)
                .where(
                    RensheExportItem.job_id == job_id,
                    RensheExportItem.application_id.in_(application_ids),
                )
                .values(status="succeeded", volume_id=volume_id, last_error=None)
            )
            job.candidate_processed += len(application_ids)
            job.volume_count = volume.volume_no
            job.heartbeat_at = now
            await db.commit()

    async def _mark_failed(self, job_id: int, exc: Exception) -> None:
        safe_error = redact_sensitive_text(
            f"{type(exc).__name__}: {str(exc)[:1000]}"
        ) or type(exc).__name__
        async with get_db_ctx() as db:
            job = await db.scalar(
                select(RensheExportJob)
                .where(RensheExportJob.id == job_id)
                .with_for_update()
            )
            if job is None or job.status == "succeeded":
                return
            now = self._now()
            job.status = "failed"
            job.finished_at = now
            job.heartbeat_at = now
            job.last_error = safe_error
            await db.execute(
                update(RensheExportVolume)
                .where(
                    RensheExportVolume.job_id == job_id,
                    RensheExportVolume.status == "running",
                )
                .values(status="failed", last_error=safe_error)
            )
            await db.execute(
                update(RensheExportItem)
                .where(
                    RensheExportItem.job_id == job_id,
                    RensheExportItem.status.in_(("queued", "running")),
                )
                .values(status="failed", last_error=safe_error)
            )
            db.add(
                RensheAuditLog(
                    actor_type="system",
                    actor_id=None,
                    action="export.generate",
                    object_type="export_job",
                    object_id=job.id,
                    result="failed",
                    summary={"error_type": type(exc).__name__},
                )
            )
            await db.commit()

    async def _recover_stale_jobs(self) -> None:
        cutoff = self._now() - timedelta(
            seconds=max(60, settings.RENSHE_WORKER_POLL_SECONDS * 12)
        )
        async with get_db_ctx() as db:
            stale_job_ids = select(RensheExportJob.id).where(
                RensheExportJob.status == "running",
                (
                    RensheExportJob.heartbeat_at.is_(None)
                    | (RensheExportJob.heartbeat_at < cutoff)
                ),
            )
            stale_error = "WorkerHeartbeatLost: 导出工作进程心跳中断"
            await db.execute(
                update(RensheExportVolume)
                .where(
                    RensheExportVolume.job_id.in_(stale_job_ids),
                    RensheExportVolume.status == "running",
                )
                .values(status="failed", last_error=stale_error)
            )
            await db.execute(
                update(RensheExportItem)
                .where(
                    RensheExportItem.job_id.in_(stale_job_ids),
                    RensheExportItem.status.in_(("queued", "running")),
                )
                .values(status="failed", last_error=stale_error)
            )
            await db.execute(
                update(RensheExportJob)
                .where(
                    RensheExportJob.status == "running",
                    (
                        RensheExportJob.heartbeat_at.is_(None)
                        | (RensheExportJob.heartbeat_at < cutoff)
                    ),
                )
                .values(
                    status="failed",
                    finished_at=self._now(),
                    last_error=stale_error,
                )
            )
            await db.commit()

    async def _heartbeat_loop(self, job_id: int) -> None:
        interval = max(5, settings.RENSHE_WORKER_POLL_SECONDS)
        while True:
            await asyncio.sleep(interval)
            async with get_db_ctx() as db:
                await db.execute(
                    update(RensheExportJob)
                    .where(
                        RensheExportJob.id == job_id,
                        RensheExportJob.status == "running",
                    )
                    .values(heartbeat_at=self._now())
                )
                await db.commit()

    async def _job_generation(self, job_id: int) -> int:
        async with get_db_ctx() as db:
            value = await db.scalar(
                select(RensheExportJob.generation_no).where(RensheExportJob.id == job_id)
            )
            if value is None:
                raise NotFoundException("人社导出任务")
            return value

    @staticmethod
    async def _eligible_candidate_rows(db, plan_id: int) -> list[tuple[int, int, int]]:
        paid_exists = exists(
            select(Order.id).where(
                Order.application_id == RensheApplication.id,
                Order.status.in_(("paid", "completed")),
            )
        )
        rows = (
            await db.execute(
                select(
                    RensheApplication.id,
                    RensheApplication.current_version_id,
                    func.coalesce(func.sum(RensheMaterial.size_bytes), 0),
                )
                .outerjoin(
                    RensheMaterial,
                    (RensheMaterial.version_id == RensheApplication.current_version_id)
                    & RensheMaterial.is_deleted.is_(False),
                )
                .where(
                    RensheApplication.plan_id == plan_id,
                    RensheApplication.status.in_(EXPORTABLE_APPLICATION_STATUSES),
                    RensheApplication.frozen_at.is_(None),
                    RensheApplication.current_version_id.is_not(None),
                    paid_exists,
                )
                .group_by(RensheApplication.id, RensheApplication.current_version_id)
                .order_by(RensheApplication.id)
            )
        ).all()
        return [(row[0], row[1], int(row[2] or 0)) for row in rows]

    @staticmethod
    def _template_paths() -> tuple[Path, Path]:
        template_dir = Path(settings.RENSHE_TEMPLATE_DIR).resolve()
        registration = template_dir / REGISTRATION_TEMPLATE_NAME
        work = template_dir / WORK_HISTORY_TEMPLATE_NAME
        if not registration.is_file() or not work.is_file():
            raise ThirdPartyException("人社 Excel 模板未配置或不存在")
        return registration, work

    @staticmethod
    def _file_size_and_sha256(path: Path) -> tuple[int, str]:
        digest = hashlib.sha256()
        size = 0
        with path.open("rb") as source:
            for chunk in iter(lambda: source.read(1024 * 1024), b""):
                size += len(chunk)
                digest.update(chunk)
        return size, digest.hexdigest()

    @staticmethod
    def _job_response(
        job: RensheExportJob, volumes: Sequence[RensheExportVolume]
    ) -> RensheExportJobResponse:
        return RensheExportJobResponse(
            id=job.id,
            plan_id=job.plan_id,
            generation_no=job.generation_no,
            requested_by_admin_id=job.requested_by_admin_id,
            status=job.status,
            candidate_total=job.candidate_total,
            candidate_processed=job.candidate_processed,
            volume_count=job.volume_count,
            heartbeat_at=job.heartbeat_at,
            started_at=job.started_at,
            finished_at=job.finished_at,
            retry_count=job.retry_count,
            last_error=job.last_error,
            volumes=[
                RensheExportVolumeResponse(
                    id=volume.id,
                    job_id=volume.job_id,
                    volume_no=volume.volume_no,
                    status=volume.status,
                    candidate_count=volume.candidate_count,
                    size_bytes=volume.size_bytes,
                    sha256=volume.sha256,
                    finished_at=volume.finished_at,
                    last_error=volume.last_error,
                    download_available=(
                        volume.status == "succeeded" and bool(volume.storage_key)
                    ),
                )
                for volume in volumes
            ],
        )

    @staticmethod
    async def _access_audit(
        *,
        actor_type: str,
        actor_id: int,
        action: str,
        object_type: str,
        object_id: int,
        application_id: int | None,
        version_id: int | None,
        material_id: int | None,
        ip_address: str | None,
        result: str,
        summary: dict,
    ) -> None:
        # A signed URL or a material upload must not be turned into a 500 just
        # because the append-only audit table is temporarily unavailable.  The
        # access event is still attempted and the failure is observable via a
        # PII-safe operational log; the database/readiness probe remains the
        # source of truth for a persistent outage.
        try:
            async with get_db_ctx() as db:
                db.add(
                    RensheAuditLog(
                        actor_type=actor_type,
                        actor_id=actor_id,
                        action=action,
                        object_type=object_type,
                        object_id=object_id,
                        application_id=application_id,
                        version_id=version_id,
                        material_id=material_id,
                        ip_address=ip_address,
                        result=result,
                        summary=summary,
                    )
                )
                await db.commit()
        except Exception:
            logger.warning(
                "renshe access audit write failed action=%s object_type=%s object_id=%s",
                action,
                object_type,
                object_id,
            )


async def renshe_export_worker_loop() -> None:
    service = RensheExportService()
    while True:
        try:
            processed = await service.process_next_job()
        except Exception as exc:
            logger.error(
                "human-resources export worker iteration failed: exception_type=%s",
                type(exc).__name__,
            )
            processed = False
        if not processed:
            await asyncio.sleep(settings.RENSHE_WORKER_POLL_SECONDS)
