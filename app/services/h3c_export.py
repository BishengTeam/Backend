"""Asynchronous official H3C workbook and image-attachment exports."""

from __future__ import annotations

import asyncio
import hashlib
import io
import logging
import shutil
import tempfile
import uuid
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path

from sqlalchemy import func, select

from app.adapter.database import get_db_ctx
from app.domain.h3c.src.index import (
    H3cExamBatch,
    H3cExportItem,
    H3cExportJob,
    H3cMaterial,
    H3cRegistration,
)
from app.domain.plan.src.index import Plan
from app.integrations.h3c_storage import H3cObjectStorage
from app.port.config import settings
from app.port.exceptions import BusinessException, ConflictException, NotFoundException, ThirdPartyException
from app.schemas.common import PaginatedData
from app.schemas.h3c_registration import (
    H3cExportCreate,
    H3cExportJobResponse,
    H3cSignedUrlResponse,
)


logger = logging.getLogger(__name__)
EXPORT_RETENTION_DAYS = 30
H3C_TEMPLATE_SHA256 = {
    "coupon": "12bea5a00028785a870220ef258261d642fab8ac1c1267c4ab44307bd1471d83",
    "full": "90238331f141e23996231ed6d2a357060b748df94cd5742353d195676a5aca80",
    "student": "c8ea41c819b0c4adfae3236f91ab5e58e072a63414a8125285ce637ccded6201",
}
MATERIAL_SLUGS = {
    "coupon_proof": "coupon-proof",
    "student_proof": "student-proof",
}


@dataclass(frozen=True, slots=True)
class _ExportRow:
    registration: H3cRegistration
    materials: tuple[H3cMaterial, ...]


class H3cExportService:
    def __init__(self, storage: H3cObjectStorage | None = None) -> None:
        self.storage = storage or H3cObjectStorage()

    async def create_job(
        self,
        *,
        admin_id: int,
        data: H3cExportCreate,
    ) -> H3cExportJobResponse:
        async with get_db_ctx() as db:
            async with db.begin():
                batch = await db.scalar(
                    select(H3cExamBatch)
                    .where(H3cExamBatch.id == data.batch_id)
                    .with_for_update()
                )
                if batch is None:
                    raise NotFoundException("H3C 考试批次")
                registrations = (
                    await db.execute(
                        select(H3cRegistration)
                        .where(
                            H3cRegistration.batch_id == batch.id,
                            H3cRegistration.registration_type == data.registration_type,
                            H3cRegistration.status.in_(data.include_statuses),
                        )
                        .order_by(H3cRegistration.id)
                        .with_for_update()
                    )
                ).scalars().all()
                job = H3cExportJob(
                    batch_id=batch.id,
                    registration_type=data.registration_type,
                    artifact_type=data.artifact_type,
                    requested_by_admin_id=admin_id,
                    include_statuses=list(data.include_statuses),
                    status="queued",
                    registration_count=len(registrations),
                )
                db.add(job)
                await db.flush()
                for registration in registrations:
                    db.add(
                        H3cExportItem(
                            job_id=job.id,
                            registration_id=registration.id,
                        )
                    )
            return await self.get_job(job.id)

    async def list_jobs(
        self,
        *,
        batch_id: int | None = None,
        status: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> PaginatedData[H3cExportJobResponse]:
        async with get_db_ctx() as db:
            stmt = select(H3cExportJob)
            if batch_id is not None:
                stmt = stmt.where(H3cExportJob.batch_id == batch_id)
            if status:
                stmt = stmt.where(H3cExportJob.status == status)
            total = await db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
            rows = (
                await db.execute(
                    stmt.order_by(H3cExportJob.id.desc())
                    .offset((page - 1) * page_size)
                    .limit(page_size)
                )
            ).scalars().all()
            return PaginatedData(
                items=[self._response(row) for row in rows],
                total=int(total),
                page=page,
                page_size=page_size,
            )

    async def get_job(self, job_id: int) -> H3cExportJobResponse:
        async with get_db_ctx() as db:
            job = await db.get(H3cExportJob, job_id)
            if job is None:
                raise NotFoundException("H3C 导出任务")
            return self._response(job)

    async def signed_url(
        self,
        *,
        admin_id: int,
            job_id: int,
    ) -> H3cSignedUrlResponse:
        async with get_db_ctx() as db:
            job = await db.get(H3cExportJob, job_id)
            if job is None:
                raise NotFoundException("H3C 导出任务")
            if job.status != "succeeded" or not job.storage_key:
                raise ConflictException("H3C 导出产物尚未生成或已过期")
            if job.expires_at is not None and job.expires_at <= datetime.now(timezone.utc):
                raise ConflictException("H3C 导出产物已过期")
            url = await self.storage.signed_get_url(job.storage_key)
            return H3cSignedUrlResponse(
                url=url,
                expires_in=settings.ALIYUN_OSS_SIGNED_URL_TTL_SECONDS,
            )

    async def process_next_job(self) -> bool:
        async with get_db_ctx() as db:
            job = await db.scalar(
                select(H3cExportJob)
                .where(H3cExportJob.status == "queued")
                .order_by(H3cExportJob.id)
                .with_for_update(skip_locked=True)
                .limit(1)
            )
            if job is None:
                return False
            job_id = job.id
            job.status = "running"
            job.started_at = datetime.now(timezone.utc)
            job.heartbeat_at = job.started_at
            await db.commit()
        try:
            await self._run_job(job_id)
            return True
        except Exception as exc:
            logger.error(
                "H3C export job failed: job_id=%s exception_type=%s",
                job_id,
                type(exc).__name__,
            )
            await self._mark_failed(job_id, exc)
            return True

    async def _run_job(self, job_id: int) -> None:
        async with get_db_ctx() as db:
            job = await db.get(H3cExportJob, job_id)
            if job is None:
                raise NotFoundException("H3C 导出任务")
            batch = await db.get(H3cExamBatch, job.batch_id)
            if batch is None:
                raise NotFoundException("H3C 考试批次")
            registration_rows = (
                await db.execute(
                    select(H3cRegistration)
                    .join(H3cExportItem, H3cExportItem.registration_id == H3cRegistration.id)
                    .where(H3cExportItem.job_id == job.id)
                    .order_by(H3cRegistration.id)
                )
            ).scalars().all()
            material_map: dict[int, tuple[H3cMaterial, ...]] = {}
            for registration in registration_rows:
                material_map[registration.id] = tuple(
                    (
                        await db.execute(
                            select(H3cMaterial)
                            .where(
                                H3cMaterial.registration_id == registration.id,
                                H3cMaterial.is_current.is_(True),
                            )
                            .order_by(H3cMaterial.material_type)
                        )
                    ).scalars().all()
                )
            rows = [
                _ExportRow(registration=row, materials=material_map[row.id])
                for row in registration_rows
            ]
            storage_key = f"h3c/exports/{batch.id}/{job.id}/{uuid.uuid4().hex}"
            if job.artifact_type == "embedded_xlsx":
                extension = "xlsx"
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                extension = "zip"
                content_type = "application/zip"
            storage_key += f".{extension}"

        with tempfile.TemporaryDirectory(prefix="wemini-h3c-export-") as temp:
            directory = Path(temp)
            if job.artifact_type == "embedded_xlsx":
                artifact = await self._build_workbook(
                    batch=batch,
                    registration_type=job.registration_type,
                    rows=rows,
                    destination=directory,
                )
            else:
                artifact = await self._build_zip(
                    rows=rows,
                    destination=directory,
                )
            await self.storage.upload_file(storage_key, artifact, content_type)
            digest = hashlib.sha256(artifact.read_bytes()).hexdigest()
            size = artifact.stat().st_size

        async with get_db_ctx() as db:
            current = await db.scalar(
                select(H3cExportJob).where(H3cExportJob.id == job_id).with_for_update()
            )
            if current is None or current.status != "running":
                raise ConflictException("H3C 导出任务状态已变化")
            now = datetime.now(timezone.utc)
            current.status = "succeeded"
            current.finished_at = now
            current.storage_key = storage_key
            current.artifact_sha256 = digest
            current.artifact_bytes = size
            current.expires_at = now + timedelta(days=EXPORT_RETENTION_DAYS)
            current.last_error = None
            await db.commit()

    async def _build_workbook(
        self,
        *,
        batch: H3cExamBatch,
        registration_type: str,
        rows: list[_ExportRow],
        destination: Path,
    ) -> Path:
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as XlsxImage
            from PIL import Image as PillowImage
        except ImportError as exc:
            raise ThirdPartyException("H3C Excel 导出依赖未安装") from exc

        template = Path(__file__).resolve().parents[1] / "templates" / "h3c" / f"{registration_type}.xlsx"
        if not template.is_file() or hashlib.sha256(template.read_bytes()).hexdigest() != H3C_TEMPLATE_SHA256[registration_type]:
            raise ThirdPartyException("H3C 官方模板缺失或已损坏")
        output = destination / f"h3c-{registration_type}-{uuid.uuid4().hex}.xlsx"
        workbook = load_workbook(template)
        worksheet = workbook["模板"]
        if worksheet.max_row >= 3:
            worksheet.delete_rows(3, worksheet.max_row - 2)

        material_dir = destination / "materials"
        material_dir.mkdir(parents=True)
        for index, row in enumerate(rows):
            excel_row = index + 3
            values = self._workbook_row(
                batch=batch,
                registration=row.registration,
                registration_type=registration_type,
            )
            for column, value in enumerate(values, 1):
                worksheet.cell(row=excel_row, column=column, value=value)
            worksheet.row_dimensions[excel_row].height = 80
            if registration_type in {"coupon", "student"}:
                material = next(
                    (
                        item
                        for item in row.materials
                        if item.material_type
                        == ("coupon_proof" if registration_type == "coupon" else "student_proof")
                    ),
                    None,
                )
                if material is not None:
                    material_path = material_dir / f"{row.registration.id}-{material.id}.jpg"
                    await self.storage.download_file(material.storage_key, material_path)
                    with PillowImage.open(material_path) as image:
                        width, height = image.size
                    scale = min(70 / width, 70 / height)
                    picture = XlsxImage(str(material_path))
                    picture.width = int(width * scale)
                    picture.height = int(height * scale)
                    worksheet.add_image(picture, f"L{excel_row}")
        workbook.save(output)
        return output

    @staticmethod
    def _workbook_row(
        *,
        batch: H3cExamBatch,
        registration: H3cRegistration,
        registration_type: str,
    ) -> list[object | None]:
        snapshot = registration.candidate_snapshot
        common = [
            snapshot.get("candidate_name"),
            snapshot.get("gender"),
            snapshot.get("candidate_idcard"),
            snapshot.get("school"),
            snapshot.get("address"),
            snapshot.get("phone"),
            snapshot.get("email"),
            batch.country,
            batch.language,
            snapshot.get("education"),
        ]
        if registration_type == "coupon":
            return [
                snapshot.get("coupon_code"),
                *common,
                None,  # image placeholder
                None,
                snapshot.get("birth_date"),
                batch.identity_tag,
                snapshot.get("first_name_en"),
                snapshot.get("last_name_en"),
                snapshot.get("exam_datetime"),
                batch.exam_code,
                None,
                batch.training_org,
                batch.training_address,
                batch.training_start,
                batch.training_end,
            ]
        if registration_type == "full":
            return [
                batch.exam_code,
                *common,
                snapshot.get("birth_date"),
                batch.identity_tag,
                snapshot.get("first_name_en"),
                snapshot.get("last_name_en"),
                snapshot.get("exam_datetime"),
                None,
                batch.training_org,
                batch.training_teacher,
                batch.training_start,
                batch.training_end,
            ]
        return [
            batch.exam_code,
            *common,
            None,  # image placeholder
            snapshot.get("verify_code"),
            snapshot.get("birth_date"),
            snapshot.get("first_name_en"),
            snapshot.get("last_name_en"),
            snapshot.get("exam_datetime"),
            None,
            batch.training_org,
            batch.training_teacher,
            batch.training_start,
            batch.training_end,
        ]

    async def _build_zip(
        self,
        *,
        rows: list[_ExportRow],
        destination: Path,
    ) -> Path:
        output = destination / f"h3c-images-{uuid.uuid4().hex}.zip"
        material_dir = destination / "zip-materials"
        material_dir.mkdir(parents=True)
        used_names: set[str] = set()
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_STORED) as archive:
            for row in rows:
                for material in row.materials:
                    slug = MATERIAL_SLUGS[material.material_type]
                    name = f"{row.registration.registration_no}-{slug}.jpg"
                    candidate = name
                    serial = 2
                    while candidate in used_names:
                        candidate = f"{name[:-4]}-{serial}.jpg"
                        serial += 1
                    used_names.add(candidate)
                    material_path = material_dir / candidate
                    await self.storage.download_file(material.storage_key, material_path)
                    archive.write(material_path, arcname=candidate)
        return output

    async def expire_artifacts(self) -> int:
        now = datetime.now(timezone.utc)
        expired_count = 0
        async with get_db_ctx() as db:
            rows = (
                await db.execute(
                    select(H3cExportJob)
                    .where(
                        H3cExportJob.status == "succeeded",
                        H3cExportJob.expires_at.is_not(None),
                        H3cExportJob.expires_at <= now,
                        H3cExportJob.storage_key.is_not(None),
                    )
                    .with_for_update(skip_locked=True)
                )
            ).scalars().all()
            for job in rows:
                if job.storage_key:
                    await self.storage.delete(job.storage_key)
                job.storage_key = None
                job.artifact_sha256 = None
                job.artifact_bytes = None
                job.expires_at = None
                expired_count += 1
            await db.commit()
        return expired_count

    async def _mark_failed(self, job_id: int, exc: Exception) -> None:
        async with get_db_ctx() as db:
            job = await db.scalar(
                select(H3cExportJob).where(H3cExportJob.id == job_id).with_for_update()
            )
            if job is None:
                return
            job.status = "failed"
            job.finished_at = datetime.now(timezone.utc)
            job.last_error = f"{type(exc).__name__}: {exc}"[:2000]
            await db.commit()

    @staticmethod
    def _response(job: H3cExportJob) -> H3cExportJobResponse:
        return H3cExportJobResponse.model_validate(job)


async def h3c_export_worker_loop(service: H3cExportService | None = None) -> None:
    active_service = service or H3cExportService()
    while True:
        try:
            processed = await active_service.process_next_job()
            if not processed:
                await active_service.expire_artifacts()
        except Exception:
            logger.warning(
                "H3C export worker iteration failed: exception_type=%s",
                type(exc).__name__,
            )
        await asyncio.sleep(5)
